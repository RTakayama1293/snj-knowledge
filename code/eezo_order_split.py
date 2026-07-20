"""
EEZO注文CSV → 仕入先別発送リスト生成スクリプト
==============================================
Shopifyのorders_export CSVを読み込み、商品名→仕入先マッピングで仕分け、
仕入先ごとのExcelファイルを出力する。

使い方:
    python eezo_order_split.py <input_csv> [--output-dir <dir>]

仕入先マッピングの更新:
    SUPPLIER_MAP の辞書を編集する。キーワード部分一致で判定。
    優先度は辞書の登録順（先にマッチしたものが採用）。
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# 仕入先マッピング定義（キーワード部分一致 → 仕入先名）
# 商品が増えたらここに追加するだけでOK
# ---------------------------------------------------------------------------
SUPPLIER_MAP = {
    # オーセントホテル小樽（ベーカリー「オンディーヌ」）
    "クッキー缶": "オーセントホテル小樽",
    "道産果実の焼き菓子": "オーセントホテル小樽",
    "マドレーヌとフィナンシェ": "オーセントホテル小樽",
    "ガレットとクッキー缶": "オーセントホテル小樽",
    "小樽のクッキー": "オーセントホテル小樽",

    # 梅屋（旭川）
    "シュークリーム": "梅屋",
    "旭川の名物": "梅屋",
    "旭川の": "梅屋",

    # トワ・ヴェール（黒松内）
    "黒松内": "トワ・ヴェール",
    "トワ・ヴェール": "トワ・ヴェール",
    "カッサータ": "トワ・ヴェール",
    "ムースフロマージュ": "トワ・ヴェール",
    "ブルーチーズケーキ": "トワ・ヴェール",

    # NIKI Hills Winery（仁木町）
    "仁木": "NIKI Hills Winery",
    "バッカス": "NIKI Hills Winery",
    "NEIRO": "NIKI Hills Winery",
    "ワイン": "NIKI Hills Winery",

    # 小樽水産加工業協同組合
    "小樽のにしん": "小樽水産加工組合",
    "小樽の酒肴": "小樽水産加工組合",
    "小樽の海鮮": "小樽水産加工組合",
    "にしんとホッケ": "小樽水産加工組合",
    "紅ずわい蟹": "小樽水産加工組合",

    # Nao-buns（倶知安）
    "倶知安": "Nao-buns",
    "Nao-buns": "Nao-buns",
}


import json as _json
from pathlib import Path as _Path

def _load_supplier_map():
    p = _Path(__file__).parent / "supplier_map.json"
    if not p.exists():
        return {}
    raw = _json.loads(p.read_text(encoding="utf-8"))
    norm = {}
    for k, v in raw.items():
        key = re.sub(r"[\s　]+", "", k)
        norm[key] = v
    return norm

_SUPPLIER_JSON = _load_supplier_map()

def detect_supplier_v06(product_name: str):
    """(仕入先, 仕入先商品名) を返す。JSON完全一致 → キーワード → 未分類"""
    key = re.sub(r"[\s　]+", "", product_name or "")
    hit = _SUPPLIER_JSON.get(key)
    if hit:
        return hit.get("supplier", "【未分類】"), hit.get("supplier_product_name", "")
    return detect_supplier(product_name), ""

def detect_supplier(product_name: str) -> str:
    """商品名からキーワード部分一致で仕入先を判定"""
    for keyword, supplier in SUPPLIER_MAP.items():
        if keyword in product_name:
            return supplier
    return "【未分類】"


# ---------------------------------------------------------------------------
# 発送に必要な列の定義
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# 電話番号正規化（国内前提）
# ---------------------------------------------------------------------------
def normalize_phone(raw: str) -> str:
    """
    +81… → 0… 変換、ハイフン挿入。
    携帯: 0X0-XXXX-XXXX / 固定(2桁局番): 0X-XXXX-XXXX / その他: そのまま
    """
    if not raw:
        return ""
    # 数字以外を除去（+は先に処理）
    s = raw.strip()
    if s.startswith("+81"):
        s = "0" + s[3:]
    s = re.sub(r"[^\d]", "", s)
    if not s:
        return raw
    # 携帯 (070/080/090)
    if re.match(r"^0[789]0", s) and len(s) == 11:
        return f"{s[:3]}-{s[3:7]}-{s[7:]}"
    # 固定 2桁局番 (03/06 等)
    if re.match(r"^0[1-9]", s) and len(s) == 10:
        return f"{s[:2]}-{s[2:6]}-{s[6:]}"
    # 050 (IP電話)
    if s.startswith("050") and len(s) == 11:
        return f"{s[:3]}-{s[3:7]}-{s[7:]}"
    # フォーマット不明はそのまま返す
    return s


# ---------------------------------------------------------------------------
# 発送に必要な列の定義
# ---------------------------------------------------------------------------
# (日本語ヘッダー, CSVカラム名, グループ)
# グループ: "order"=注文情報, "billing"=依頼主, "shipping"=送り先, "other"=その他
OUTPUT_COLUMNS = [
    # --- 注文情報 ---
    ("注文番号",        "Name",                   "order"),
    ("注文日",          "Created at",             "order"),
    ("商品名",          "Lineitem name",          "order"),
    ("仕入先商品名",    "__supplier_product__",   "order"),
    ("数量",            "Lineitem quantity",      "order"),
    # --- 依頼主（請求先） ---
    ("依頼主",          "Billing Name",           "billing"),
    ("依頼主TEL",       "Billing Phone",          "billing"),
    ("依頼主〒",        "Billing Zip",            "billing"),
    ("依頼主 都道府県", "Billing Province Name",  "billing"),
    ("依頼主 市区町村", "Billing City",           "billing"),
    ("依頼主 住所1",    "Billing Address1",       "billing"),
    ("依頼主 住所2",    "Billing Address2",       "billing"),
    # --- 送り先（送付先） ---
    ("送り先",          "Shipping Name",          "shipping"),
    ("送り先TEL",       "Shipping Phone",         "shipping"),
    ("送り先〒",        "Shipping Zip",           "shipping"),
    ("送り先 都道府県", "Shipping Province Name", "shipping"),
    ("送り先 市区町村", "Shipping City",          "shipping"),
    ("送り先 住所1",    "Shipping Address1",      "shipping"),
    ("送り先 住所2",    "Shipping Address2",      "shipping"),
    # --- その他 ---
    ("備考",            "Notes",                  "other"),
]

# 電話番号カラム（正規化対象）
PHONE_COLUMNS = {"依頼主TEL", "送り先TEL"}


def parse_orders(csv_path: str) -> list[dict]:
    """ShopifyのCSVを読み込み、発送用の行リストを返す"""
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            row = {}
            for ja_name, en_key, _ in OUTPUT_COLUMNS:
                val = raw.get(en_key, "").strip()
                if en_key == "Created at" and val:
                    val = val[:10]
                if ja_name in PHONE_COLUMNS:
                    val = normalize_phone(val)
                row[ja_name] = val
            supplier, sp_name = detect_supplier_v06(raw.get("Lineitem name", ""))
            row["仕入先"] = supplier
            row["仕入先商品名"] = sp_name
            rows.append(row)

    # 宛先の前方補完（複数明細注文は2行目以降の宛先が空になるShopify仕様に対応）
    fill_cols = [c for c, _, g in OUTPUT_COLUMNS if g in ("billing", "shipping", "other")]
    first_by_order = {}
    for r in rows:
        name = r.get("注文番号", "")
        if name not in first_by_order:
            first_by_order[name] = r
        else:
            base = first_by_order[name]
            for c in fill_cols:
                if not r.get(c):
                    r[c] = base.get(c, "")
    return rows


def write_supplier_xlsx(supplier: str, orders: list[dict], output_dir: Path, date_str: str):
    """仕入先1社分のExcelファイルを出力"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FONT = Font(name='Meiryo UI', size=10)
    FONT_BOLD = Font(name='Meiryo UI', size=10, bold=True)
    FONT_TITLE = Font(name='Meiryo UI', size=14, bold=True)
    FONT_GROUP = Font(name='Meiryo UI', size=9, bold=True)
    FILL_HEADER = PatternFill('solid', fgColor='D9D9D9')
    FILL_BILLING = PatternFill('solid', fgColor='E8F0FE')   # 薄青 — 依頼主
    FILL_SHIPPING = PatternFill('solid', fgColor='FFF2CC')   # 薄黄 — 送り先
    THIN_BORDER = Border(bottom=Side(style='thin', color='000000'))
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_WRAP = Alignment(wrap_text=True, vertical='top')

    wb = Workbook()
    ws = wb.active
    ws.title = "発送リスト"

    safe_name = supplier.replace("/", "／")
    num_cols = len(OUTPUT_COLUMNS)

    # --- Row 1: タイトル ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws["A1"]
    title_cell.value = f"EEZO 発送依頼書 — {safe_name}"
    title_cell.font = FONT_TITLE
    title_cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 30

    # --- Row 2: メタ情報 ---
    ws["A2"].value = f"出力日: {date_str}　件数: {len(orders)}件"
    ws["A2"].font = Font(name='Meiryo UI', size=9, italic=True)

    # --- Row 3: グループヘッダー（依頼主 / 送り先） ---
    group_row = 3
    groups = {"order": "注文情報", "billing": "依頼主（請求先）", "shipping": "送り先（送付先）", "other": ""}
    group_fills = {"billing": FILL_BILLING, "shipping": FILL_SHIPPING}

    # グループごとの開始・終了列を計算
    group_ranges = {}
    for col_idx, (_, _, grp) in enumerate(OUTPUT_COLUMNS, 1):
        if grp not in group_ranges:
            group_ranges[grp] = [col_idx, col_idx]
        else:
            group_ranges[grp][1] = col_idx

    for grp, (start_col, end_col) in group_ranges.items():
        label = groups.get(grp, "")
        if not label:
            continue
        if end_col > start_col:
            ws.merge_cells(start_row=group_row, start_column=start_col, end_row=group_row, end_column=end_col)
        cell = ws.cell(row=group_row, column=start_col, value=label)
        cell.font = FONT_GROUP
        cell.alignment = ALIGN_CENTER
        fill = group_fills.get(grp)
        if fill:
            for c in range(start_col, end_col + 1):
                ws.cell(row=group_row, column=c).fill = fill

    # --- Row 4: カラムヘッダー ---
    header_row = 4
    for col_idx, (ja, _, grp) in enumerate(OUTPUT_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=ja)
        cell.font = FONT_BOLD
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # --- Row 5+: データ ---
    wrap_cols = {"依頼主 住所1", "依頼主 住所2", "送り先 住所1", "送り先 住所2", "備考"}
    center_cols = {"数量", "依頼主〒", "送り先〒"}
    for row_idx, order in enumerate(orders, header_row + 1):
        for col_idx, (ja, _, grp) in enumerate(OUTPUT_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=order[ja])
            cell.font = FONT
            if ja in center_cols:
                cell.alignment = ALIGN_CENTER
            elif ja in wrap_cols:
                cell.alignment = ALIGN_WRAP
            else:
                cell.alignment = Alignment(vertical='top')
            # グループの背景を薄く引く
            fill = group_fills.get(grp)
            if fill:
                cell.fill = fill

    # --- 列幅 ---
    col_widths = {
        "注文番号": 14, "注文日": 12, "商品名": 28, "数量": 5,
        "依頼主": 12, "依頼主TEL": 15, "依頼主〒": 10,
        "依頼主 都道府県": 9, "依頼主 市区町村": 10,
        "依頼主 住所1": 22, "依頼主 住所2": 18,
        "送り先": 12, "送り先TEL": 15, "送り先〒": 10,
        "送り先 都道府県": 9, "送り先 市区町村": 10,
        "送り先 住所1": 22, "送り先 住所2": 18,
        "備考": 14,
    }
    for col_idx, (ja, _, _) in enumerate(OUTPUT_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(ja, 12)

    # --- 印刷設定 ---
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    filename = f"{date_str}_発送リスト_{safe_name}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    return filepath


def main():
    parser = argparse.ArgumentParser(description="EEZO注文CSV → 仕入先別発送リスト")
    parser.add_argument("input_csv", help="Shopify orders_export CSV")
    parser.add_argument("--output-dir", default=".", help="出力先ディレクトリ")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    date_str = datetime.now().strftime("%Y%m%d")
    orders = parse_orders(args.input_csv)

    # 仕入先別に仕分け
    by_supplier = defaultdict(list)
    for o in orders:
        by_supplier[o["仕入先"]].append(o)

    # サマリー表示
    print(f"\n📦 注文合計: {len(orders)}件 / 仕入先: {len(by_supplier)}社\n")
    print(f"{'仕入先':<20} {'件数':>4}  商品")
    print("-" * 70)
    for supplier, items in sorted(by_supplier.items()):
        products = ", ".join(set(i["商品名"] for i in items))
        print(f"{supplier:<20} {len(items):>4}  {products}")
    print()

    # Excel出力
    created_files = []
    for supplier, items in sorted(by_supplier.items()):
        fp = write_supplier_xlsx(supplier, items, output_dir, date_str)
        created_files.append(fp)
        print(f"✅ {fp.name}")

    # 全件まとめ版も出力
    fp_all = write_supplier_xlsx("全仕入先", orders, output_dir, date_str)
    created_files.append(fp_all)
    print(f"✅ {fp_all.name}（全件まとめ）")

    return created_files


if __name__ == "__main__":
    main()
